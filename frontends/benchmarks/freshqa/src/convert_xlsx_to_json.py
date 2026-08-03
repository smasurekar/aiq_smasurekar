#!/usr/bin/env python3
"""Convert the official FreshQA XLSX release to the JSON format NAT evaluation expects.

The upstream FreshQA release (https://github.com/freshllms/freshqa) ships as an
``.xlsx`` workbook rather than a CSV, so :mod:`convert_csv_to_json` cannot read it
directly. Two quirks of that workbook are handled here:

* The first sheet (``freshqa``) starts with a warning banner and a blank row, so the
  real header row is not row 1. We locate it by searching for the ``question`` column.
* Answers live in ``answer_0`` .. ``answer_9`` columns, only some of which are filled.

The emitted records carry both the fields the eval dataset reader needs
(``question`` / ``expected_output``) and the metadata columns the FreshQA evaluator
uses for its accuracy breakdowns (``split``, ``fact_type``, ``num_hops``,
``false_premise``, ``effective_year``).

Usage::

    python convert_xlsx_to_json.py IN.xlsx OUT.json [--split TEST] [--limit 10]
    python convert_xlsx_to_json.py IN.xlsx OUT.json [--sample 500] [--seed 42]
"""

import argparse
import json
import random
from pathlib import Path
from typing import Any

import openpyxl

# Number of answer_N columns defined by the FreshQA schema.
MAX_ANSWERS = 10


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Return the index of the header row.

    The workbook prepends a warning banner and a blank row, so we cannot assume row 0.
    The header is identified as the first row containing a ``question`` cell.
    """
    for idx, row in enumerate(rows):
        cells = {str(c).strip().lower() for c in row if c is not None}
        if "question" in cells:
            return idx
    raise ValueError("Could not locate a header row containing a 'question' column")


def convert_xlsx_to_json(
    input_xlsx: str,
    output_json: str,
    sheet_name: str = "freshqa",
    split_filter: str | None = None,
    limit: int | None = None,
    sample: int | None = None,
    seed: int = 42,
) -> int:
    """Convert a FreshQA workbook to the NAT evaluation JSON format.

    Args:
        input_xlsx: Path to the downloaded ``FreshQA_*.xlsx`` workbook.
        output_json: Path of the JSON file to write.
        sheet_name: Worksheet holding the questions (the release names it ``freshqa``).
        split_filter: Optional split to keep, e.g. ``"TEST"`` or ``"DEV"``.
        limit: Optional cap on the number of records written, taking the first N rows in
            workbook order. Cheap and reproducible, but the workbook groups rows by
            category, so use it only for small smoke sets.
        sample: Optional random subset size. Unlike ``limit`` this draws from the whole
            filtered set, so the fact-type / hop-count / premise mix stays representative.
            Preferred for any subset large enough to report accuracy on.
        seed: RNG seed for ``sample``, so the same subset is reproducible across runs.

    Returns:
        The number of records written.
    """
    if limit is not None and sample is not None:
        raise ValueError("Use either --limit or --sample, not both")
    workbook = openpyxl.load_workbook(input_xlsx, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col = {name: i for i, name in enumerate(header) if name}

    def cell(row: tuple[Any, ...], name: str) -> Any:
        """Read a named column from a row, tolerating short/ragged rows."""
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        question = cell(row, "question")
        if question is None or not str(question).strip():
            continue  # trailing blank rows

        split = cell(row, "split")
        if split_filter and str(split).strip().upper() != split_filter.upper():
            continue

        # Collect only the populated answer_N columns; the evaluator joins them with " | ".
        answers = [
            str(cell(row, f"answer_{i}")).strip()
            for i in range(MAX_ANSWERS)
            if cell(row, f"answer_{i}") is not None and str(cell(row, f"answer_{i}")).strip()
        ]

        # ids arrive from openpyxl as floats (0.0, 1.0, ...); normalize to ints so they
        # match the string keys the evaluator builds when joining metadata to results.
        raw_id = cell(row, "id")
        item_id = int(raw_id) if isinstance(raw_id, float) and raw_id.is_integer() else raw_id
        if item_id is None:
            item_id = len(records)

        record: dict[str, Any] = {
            "id": item_id,
            "question": str(question).strip(),
            "expected_output": {f"answer_{i}": answer for i, answer in enumerate(answers)},
        }

        # Metadata columns drive the evaluator's per-dimension accuracy breakdowns.
        for field in ("split", "fact_type", "num_hops", "effective_year"):
            value = cell(row, field)
            if value is not None and str(value).strip():
                record[field] = str(value).strip()

        false_premise = cell(row, "false_premise")
        if false_premise is not None:
            record["false_premise"] = bool(false_premise)

        records.append(record)
        if limit is not None and len(records) >= limit:
            break

    if sample is not None and sample < len(records):
        # Sample from the full filtered set, then restore workbook order so the output
        # file stays easy to diff and eyeball against the source spreadsheet.
        chosen = random.Random(seed).sample(range(len(records)), sample)
        records = [records[i] for i in sorted(chosen)]

    output_path = Path(output_json)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(records)} records to {output_json}")
    return len(records)


def main() -> None:
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Convert a FreshQA XLSX release to NAT evaluation JSON.")
    parser.add_argument("input_xlsx", help="Path to FreshQA_*.xlsx")
    parser.add_argument("output_json", help="Path of the JSON file to write")
    parser.add_argument("--sheet-name", default="freshqa", help="Worksheet name (default: freshqa)")
    parser.add_argument("--split", default=None, help="Keep only this split, e.g. TEST or DEV")
    parser.add_argument("--limit", type=int, default=None, help="Write the first N records (workbook order)")
    parser.add_argument("--sample", type=int, default=None, help="Write a random N-record subset (representative)")
    parser.add_argument("--seed", type=int, default=42, help="RNG seed for --sample (default: 42)")
    args = parser.parse_args()

    convert_xlsx_to_json(
        input_xlsx=args.input_xlsx,
        output_json=args.output_json,
        sheet_name=args.sheet_name,
        split_filter=args.split,
        limit=args.limit,
        sample=args.sample,
        seed=args.seed,
    )


if __name__ == "__main__":
    main()
